import argparse
import csv
import html
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

from openpyxl import load_workbook
from pypdf import PdfReader


OPERATION_WORDS = [
    "transfert de garantie out",
    "transfert de garantie in",
    "transfert d'actifs out",
    "transfert d'actifs in",
    "transfert d'actifs",
    "transfert d\u2019actifs out",
    "transfert d\u2019actifs in",
    "transfert d\u2019actifs",
    "exercice de droits donnant acces au capital",
    "exercice de droits donnant acces",
    "pret - mise en place",
    "pr\u00eat - mise en place",
    "pr\u00eat \u2013 mise en place",
    "emprunt - mise en place",
    "emprunt \u2013 mise en place",
    "pret - retour",
    "pr\u00eat - retour",
    "pr\u00eat \u2013 retour",
    "emprunt - retour",
    "emprunt \u2013 retour",
    "reduction d'une position longue",
    "accroissement d'une position longue",
    "reduction d'une position courte",
    "accroissement d'une position courte",
    "acquisition",
    "cession",
    "souscription",
    "conversion",
    "apport",
    "achat",
    "vente",
]

LEGAL_SUFFIX_FRAGMENTS = {
    "SA",
    "SAS",
    "SASU",
    "SCA",
    "SE",
    "SARL",
    "PLC",
    "LTD",
    "LLP",
    "LLC",
    "LP",
    "INC",
    "CO",
    "CORP",
    "LIMITED",
    "GMBH",
    "AG",
    "NV",
    "BV",
    "SPA",
}

CONTINUATION_END_WORDS = {
    "AND",
    "DE",
    "DES",
    "DU",
    "ET",
    "LA",
    "LE",
    "LES",
    "OF",
    "THE",
}


def strip_accents(value):
    return "".join(
        char
        for char in unicodedata.normalize("NFD", value)
        if unicodedata.category(char) != "Mn"
    )


def normalise(value):
    return re.sub(r"\s+", " ", strip_accents(value).lower()).strip()


def clean_text(value):
    return re.sub(r"\s+", " ", (value or "").strip())


def extract_pdf_text(path):
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def parse_date_fr(value):
    months = {
        "janvier": "01",
        "fevrier": "02",
        "mars": "03",
        "avril": "04",
        "mai": "05",
        "juin": "06",
        "juillet": "07",
        "aout": "08",
        "septembre": "09",
        "octobre": "10",
        "novembre": "11",
        "decembre": "12",
    }
    text = normalise(value)
    match = re.search(r"\b(\d{1,2})(?:er)?\s+([a-z]+)\s+(\d{4})\b", text)
    if not match:
        return None
    day, month_name, year = match.groups()
    month = months.get(month_name)
    if not month:
        return None
    return f"{year}-{month}-{int(day):02d}"


def parse_ddmmyyyy(value):
    match = re.search(r"\b(\d{2})/(\d{2})/(\d{4})\b", value)
    if not match:
        return None
    day, month, year = match.groups()
    return f"{year}-{month}-{day}"


def operation_index(line):
    low = normalise(line).replace("\u2019", "'")
    best = None
    for word in OPERATION_WORDS:
        idx = low.find(normalise(word).replace("\u2019", "'"))
        if idx >= 0 and (best is None or idx < best):
            best = idx
    return best


def clean_filer(value):
    return re.sub(r"[*]+$", "", clean_text(value)).strip()


def company_fragment_key(value):
    return re.sub(r"[^A-Z0-9]+", "", strip_accents(clean_filer(value)).upper())


def company_fragment_tokens(value):
    return re.findall(r"[A-Z0-9]+", strip_accents(clean_filer(value)).upper())


def is_legal_suffix_fragment(value):
    key = company_fragment_key(value)
    if not key:
        return False
    if key in LEGAL_SUFFIX_FRAGMENTS:
        return True
    tokens = company_fragment_tokens(value)
    return bool(tokens) and all(token in LEGAL_SUFFIX_FRAGMENTS for token in tokens)


def should_join_filer_fragment(previous, prefix):
    previous = clean_filer(previous)
    prefix = clean_filer(prefix)
    if not previous or not prefix:
        return False
    if is_legal_suffix_fragment(prefix):
        return True
    if previous.rstrip().endswith((",", "&", "-", "\u2013", "'", "\u2019")):
        return True
    tokens = company_fragment_tokens(previous)
    return bool(tokens) and tokens[-1] in CONTINUATION_END_WORDS


def merge_filer_fragment(previous, prefix):
    previous = clean_filer(previous)
    prefix = clean_filer(prefix)
    if not previous:
        return prefix
    if not prefix:
        return previous
    separator = "" if previous.rstrip().endswith(("&", "-", "\u2013", "'", "\u2019")) else " "
    return clean_filer(f"{previous}{separator}{prefix}")


def split_operation_line(line, current_filer):
    idx = operation_index(line)
    if idx is None:
        return None

    prefix = clean_text(line[:idx])
    rest = clean_text(line[idx:])
    date_match = re.search(r"\b(?:le|du)\s+(\d{2}/\d{2}/\d{4})\b", rest, flags=re.I)
    if not date_match:
        return None

    filer_prefix = clean_filer(prefix) if prefix else None
    if filer_prefix and current_filer and should_join_filer_fragment(current_filer, filer_prefix):
        filer = merge_filer_fragment(current_filer, filer_prefix)
    else:
        filer = filer_prefix or current_filer
    operation = clean_text(rest[: date_match.end()])
    after = clean_text(rest[date_match.end() :])

    quantity = after if is_quantity_line(after) else None

    return {
        "filer": filer,
        "operation": operation,
        "transaction_date": parse_ddmmyyyy(date_match.group(1)),
        "inline_quantity": quantity,
    }


def is_quantity_line(line):
    return bool(
        re.match(
            r"^[+-]?\d[\d .,\u00a0]*\s+(actions?|droits?|titres?|obligations?|oceanes?|oceane|bons?|equity\s+swaps?|swaps?|cfds?|cfd)\b",
            clean_text(line),
            flags=re.I,
        )
    )


def parse_quantity(line):
    text = clean_text(line).replace("\u00a0", " ")
    match = re.match(r"^([+-]?\d[\d .,\u00a0]*)(?:\s+)(.+)$", text)
    if not match:
        return {"quantity": None, "security": text}

    raw_number, security = match.groups()
    number = raw_number.replace(" ", "").replace("\u00a0", "").replace(".", "")
    try:
        quantity = int(number)
    except ValueError:
        quantity = None

    return {
        "quantity": quantity,
        "quantity_text": raw_number.strip(),
        "security": clean_text(security),
    }


def parse_price_and_holdings(lines, expected_prices=None):
    prices = []
    holdings = []

    for raw in lines:
        line = clean_text(raw)
        if not line:
            continue
        if line.startswith("*") or set(line) <= {"_"}:
            break

        negative_holding = re.match(r"^-\s+(\d[\d .,\u00a0]*(?:.*)?)$", line)
        if negative_holding and expected_prices is not None and len(prices) >= expected_prices:
            holdings.append(f"- {clean_text(negative_holding.group(1))}")
            continue

        match = re.match(r"^(-|\d+[,.]\d+)(?:\s+(.+))?$", line)
        if match:
            price, rest = match.groups()
            prices.append(price)
            if rest:
                holdings.append(clean_text(rest))
            continue

        holdings.append(line)

    return prices, holdings


def is_code_line(line):
    return bool(re.search(r"\bcode\s+FR[A-Z0-9]{10}\b", line, flags=re.I))


def is_footnote_line(line):
    text = clean_text(line)
    return text.startswith("*") or bool(text) and set(text) <= {"_"}


def is_price_line(line):
    return bool(re.match(r"^(-|\d+[,.]\d+)(?:\s+.+)?$", clean_text(line)))


def is_probable_filer_line(line):
    text = clean_text(line)
    lowered = normalise(text)
    if not text or is_footnote_line(text):
        return False
    if re.search(r"\d{2}/\d{2}/\d{4}", text):
        return False
    if is_quantity_line(text) or is_price_line(text):
        return False
    if "actions et" in lowered or "droits de vote" in lowered:
        return False
    return bool(re.search(r"[A-Za-z]", text))


def find_section_boundary(candidate):
    for idx, line in enumerate(candidate):
        if is_footnote_line(line):
            return idx
        if idx == 0:
            continue
        if (
            idx > 0
            and idx + 1 < len(candidate)
            and is_probable_filer_line(candidate[idx - 1])
            and is_legal_suffix_fragment(line)
        ):
            joined_filer = merge_filer_fragment(candidate[idx - 1], line)
            parsed_next = split_operation_line(candidate[idx + 1], joined_filer)
            if parsed_next:
                return idx - 1
        if is_probable_filer_line(line) and idx + 1 < len(candidate):
            parsed_next = split_operation_line(candidate[idx + 1], clean_filer(line))
            if parsed_next:
                return idx
        if split_operation_line(line, None):
            return idx
    return None


def make_table_segments(body):
    segments = []
    cursor = 0

    while cursor < len(body):
        code_idx = next((idx for idx in range(cursor, len(body)) if is_code_line(body[idx])), None)
        if code_idx is None:
            break

        next_code_idx = next(
            (idx for idx in range(code_idx + 1, len(body)) if is_code_line(body[idx])),
            None,
        )

        if next_code_idx is None:
            post_code = body[code_idx + 1 :]
            segments.append((body[cursor:code_idx], body[code_idx], post_code))
            break

        candidate = body[code_idx + 1 : next_code_idx]
        boundary = find_section_boundary(candidate)
        if boundary is None:
            boundary = len(candidate)

        post_code = candidate[:boundary]
        segments.append((body[cursor:code_idx], body[code_idx], post_code))
        cursor = code_idx + 1 + boundary

    return segments


def classify_operation(operation):
    text = normalise(operation)
    if "achat" in text or "acquisition" in text or "souscription" in text:
        return "buy"
    if "vente" in text or "cession" in text or "apport" in text:
        return "sell"
    if "out" in text:
        return "transfer_out"
    if "in" in text:
        return "transfer_in"
    if ("pret" in text or "emprunt" in text) and "mise en place" in text:
        return "loan_setup"
    if "pret" in text or "emprunt" in text:
        return "loan_return"
    if "position longue" in text or "position courte" in text:
        return "derivative_position"
    if "exercice" in text:
        return "exercise"
    return "other"


def classify_instrument(security):
    text = normalise(security or "")
    if any(term in text for term in ["equity swap", "swap", "cfd", "contract for difference"]):
        return "Derivative"
    if any(term in text for term in ["action", "droit de vote"]):
        return "Shares"
    if any(term in text for term in ["oceane", "obligation", "bon"]):
        return "Derivative"
    return "Other" if text else None


def split_appended_filer(holding):
    match = re.search(r"\*([A-Z][A-Za-z0-9 .,&'’()-]{2,})$", holding or "")
    if not match:
        return holding, None

    candidate = clean_filer(match.group(1))
    if re.search(r"\d", candidate):
        return holding, None

    cleaned_holding = clean_text(holding[: match.start()] + "*")
    return cleaned_holding, candidate


def parse_table_segment(record, pre_code, code_line, post_code, fallback_filer=None):
    target = record.get("target", {}).get("name")
    isin_match = re.search(r"\bFR[A-Z0-9]{10}\b", code_line)
    isin = isin_match.group(0) if isin_match else None

    current_filer = fallback_filer
    operations = []
    quantities = []

    for line in pre_code:
        if is_quantity_line(line):
            quantities.append(line)
            continue

        parsed = split_operation_line(line, current_filer)
        if parsed:
            current_filer = parsed["filer"] or current_filer
            operations.append(parsed)
            if parsed["inline_quantity"]:
                quantities.append(parsed["inline_quantity"])
            continue

        if not re.search(r"\d{2}/\d{2}/\d{4}", line):
            filer_line = clean_filer(line)
            if current_filer and should_join_filer_fragment(current_filer, filer_line):
                current_filer = merge_filer_fragment(current_filer, filer_line)
            elif current_filer is None:
                current_filer = filer_line

    prices, holdings = parse_price_and_holdings(post_code, expected_prices=len(operations))
    appended_filers = []
    cleaned_holdings = []
    for holding in holdings:
        cleaned_holding, appended_filer = split_appended_filer(holding)
        cleaned_holdings.append(cleaned_holding)
        if appended_filer:
            appended_filers.append(appended_filer)
    holdings = cleaned_holdings
    if current_filer is None and appended_filers:
        current_filer = appended_filers[0]

    if len(prices) > len(operations):
        extra = len(prices) - len(operations)
        adjusted_prices = []
        for price in prices:
            if extra and price == "-":
                extra -= 1
                continue
            adjusted_prices.append(price)
        prices = adjusted_prices
    rows = []
    notes = []

    if len(operations) != len(quantities):
        notes.append(f"Operation count {len(operations)} differs from quantity count {len(quantities)}.")
    if prices and len(prices) != len(operations):
        notes.append(f"Operation count {len(operations)} differs from price count {len(prices)}.")
    if holdings and len(holdings) not in {1, len(operations)}:
        notes.append(
            f"Found {len(holdings)} resulting holding lines for {len(operations)} operations; kept as notice-level holdings."
        )

    for idx, op in enumerate(operations):
        quantity_info = parse_quantity(quantities[idx]) if idx < len(quantities) else {}
        resulting_holding = None
        if len(holdings) == len(operations):
            resulting_holding = holdings[idx]
        elif len(holdings) == 1 and idx == len(operations) - 1:
            resulting_holding = holdings[0]

        rows.append(
            {
                "amf_number": record.get("amf_number"),
                "target": target,
                "filer": op.get("filer") or current_filer,
                "operation": op.get("operation"),
                "operation_type": classify_operation(op.get("operation") or ""),
                "transaction_date": op.get("transaction_date"),
                "published_at": record.get("published_at"),
                "online_at": record.get("online_at"),
                "date_information": record.get("date_information"),
                "isin": isin,
                "instrument_type": classify_instrument(quantity_info.get("security")),
                "quantity": quantity_info.get("quantity"),
                "quantity_text": quantity_info.get("quantity_text"),
                "security": quantity_info.get("security"),
                "price_eur": prices[idx] if idx < len(prices) else None,
                "resulting_holding": resulting_holding,
                "document_url": record.get("document", {}).get("url"),
                "document_file": record.get("document", {}).get("local_pdf"),
            }
        )

    if not operations:
        notes.append("No dated operation lines were extracted.")

    status = "ok" if rows and not notes else "review" if rows else "failed"
    return rows, status, notes, holdings


def parse_statement(record, text):
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    doc_date = parse_date_fr(lines[3]) if len(lines) > 3 else None

    start = None
    for i, line in enumerate(lines):
        nearby_header = lines[max(0, i - 8) : i + 1]
        if "transaction" in normalise(line) and any(
            "operateur" in normalise(header_line) for header_line in nearby_header
        ):
            start = i + 1
            break

    if start is None:
        return [], {
            "document_date": doc_date,
            "parse_status": "review",
            "parse_notes": ["Could not find the AMF table header."],
            "raw_text": text,
        }

    body = lines[start:]
    segments = make_table_segments(body)

    if not segments:
        return [], {
            "document_date": doc_date,
            "parse_status": "review",
            "parse_notes": ["Could not find a securities code line."],
            "raw_text": text,
        }

    rows = []
    notes = []
    holdings = []
    statuses = []

    for idx, (pre_code, code_line, post_code) in enumerate(segments):
        fallback_filer = None
        if idx + 1 < len(segments):
            next_pre = segments[idx + 1][0]
            if next_pre and is_probable_filer_line(next_pre[0]):
                fallback_filer = clean_filer(next_pre[0])
                if len(next_pre) > 1 and is_legal_suffix_fragment(next_pre[1]):
                    fallback_filer = merge_filer_fragment(fallback_filer, next_pre[1])

        segment_rows, segment_status, segment_notes, segment_holdings = parse_table_segment(
            record,
            pre_code,
            code_line,
            post_code,
            fallback_filer=fallback_filer,
        )
        rows.extend(segment_rows)
        notes.extend(segment_notes)
        holdings.extend(segment_holdings)
        statuses.append(segment_status)

    if not rows:
        notes.append("No dated operation lines were extracted.")

    status = "ok" if rows and not notes else "review" if rows else "failed"
    known_filers = {row.get("filer") for row in rows if row.get("filer")}
    if len(known_filers) == 1:
        only_filer = next(iter(known_filers))
        for row in rows:
            if not row.get("filer"):
                row["filer"] = only_filer
    return rows, {
        "document_date": doc_date,
        "parse_status": status,
        "parse_notes": notes,
        "notice_level_resulting_holdings": holdings,
        "raw_text": text,
    }


def parse_offer_company_workbook(path):
    if not path or not Path(path).exists():
        return None

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    companies = []
    last_update = None

    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        for idx, value in enumerate(values):
            if isinstance(value, str) and "last update" in normalise(value):
                for candidate in values[idx + 1 :]:
                    if isinstance(candidate, datetime):
                        last_update = candidate.date().isoformat()
                        break
            if last_update:
                break
        if last_update:
            break

    for row in sheet.iter_rows(min_row=1, values_only=True):
        offeree = row[1] if len(row) > 1 else None
        isin = row[2] if len(row) > 2 else None
        offeror = row[4] if len(row) > 4 else None
        announced = row[7] if len(row) > 7 else None
        filed = row[8] if len(row) > 8 else None
        if not offeree or str(offeree).startswith(("Notes", "1.", "2.", "3.", "4.")):
            continue
        if "SOCIETE VISEE" in str(offeree):
            continue
        if not isin:
            continue
        companies.append(
            {
                "offeree": clean_text(str(offeree)),
                "isin": clean_text(str(isin)),
                "offeror": clean_text(str(offeror)) if offeror else None,
                "offer_announced": announced.date().isoformat()
                if isinstance(announced, datetime)
                else None,
                "draft_offer_filed": filed.date().isoformat() if isinstance(filed, datetime) else None,
            }
        )

    return {"last_update": last_update, "companies": companies}


def write_csv(path, rows):
    fields = [
        "amf_number",
        "target",
        "filer",
        "operation_type",
        "operation",
        "transaction_date",
        "published_at",
        "date_information",
        "isin",
        "instrument_type",
        "quantity",
        "security",
        "price_eur",
        "resulting_holding",
        "document_url",
    ]
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_rss(path, filings, rows):
    items = []
    by_notice = {}
    for row in rows:
        by_notice.setdefault(row["amf_number"], []).append(row)

    for filing in filings:
        number = filing["amf_number"]
        notice_rows = by_notice.get(number, [])
        first = notice_rows[0] if notice_rows else {}
        target = filing.get("target", {}).get("name") or "Unknown target"
        filer = first.get("filer") or "Unknown filer"
        title = f"{target}: {filer} ({number})"
        description = f"{len(notice_rows)} parsed transaction row(s)."
        link = filing.get("document", {}).get("url") or ""
        pub_date = filing.get("date_information") or filing.get("online_at") or filing.get("published_at") or ""
        items.append(
            f"""    <item>
      <title>{xml_escape(title)}</title>
      <link>{xml_escape(link)}</link>
      <guid>{xml_escape(number)}</guid>
      <pubDate>{xml_escape(pub_date)}</pubDate>
      <description>{xml_escape(description)}</description>
    </item>"""
        )

    content = f"""<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0">
  <channel>
    <title>French OPA dealings virtual feed</title>
    <link>https://bdif.amf-france.org/en?typesInformation=OPA</link>
    <description>Proof-of-concept feed for AMF DeclarationAchatVente filings.</description>
{chr(10).join(items)}
  </channel>
</rss>
"""
    path.write_text(content, encoding="utf-8")


def write_html(path, rows, summaries, source):
    row_json = json.dumps(rows, ensure_ascii=False)
    ok = sum(1 for summary in summaries if summary["parse_status"] == "ok")
    review = sum(1 for summary in summaries if summary["parse_status"] == "review")

    content = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>French OPA Dealings Tracker POC</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #172026;
      --muted: #5b6670;
      --line: #d8dee4;
      --surface: #ffffff;
      --band: #f5f7f8;
      --blue: #1f6feb;
      --green: #197b48;
      --amber: #a15c00;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background: var(--surface);
      line-height: 1.4;
    }}
    header {{
      padding: 20px 28px 16px;
      border-bottom: 1px solid var(--line);
      background: var(--band);
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 24px;
      letter-spacing: 0;
    }}
    .meta {{
      display: flex;
      gap: 18px;
      flex-wrap: wrap;
      color: var(--muted);
      font-size: 13px;
    }}
    main {{ padding: 18px 28px 28px; }}
    .toolbar {{
      display: grid;
      grid-template-columns: minmax(180px, 260px) minmax(180px, 260px) auto;
      gap: 10px;
      align-items: end;
      margin-bottom: 14px;
    }}
    label {{
      display: grid;
      gap: 4px;
      font-size: 12px;
      color: var(--muted);
    }}
    input, select {{
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 4px;
      padding: 6px 8px;
      font: inherit;
      background: #fff;
      color: var(--ink);
    }}
    button {{
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 4px;
      padding: 6px 10px;
      background: #fff;
      color: var(--ink);
      cursor: pointer;
    }}
    .count {{ color: var(--muted); font-size: 13px; }}
    table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      font-size: 13px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 8px 7px;
      vertical-align: top;
      overflow-wrap: anywhere;
    }}
    th {{
      text-align: left;
      color: var(--muted);
      font-size: 11px;
      text-transform: uppercase;
      background: #fbfbfc;
      position: sticky;
      top: 0;
    }}
    .type-buy {{ color: var(--green); font-weight: 700; }}
    .type-sell {{ color: #b42318; font-weight: 700; }}
    .type-transfer_in, .type-transfer_out, .type-loan_return, .type-loan_setup {{ color: var(--amber); font-weight: 700; }}
    a {{ color: var(--blue); }}
    @media (max-width: 760px) {{
      header, main {{ padding-left: 14px; padding-right: 14px; }}
      .toolbar {{ grid-template-columns: 1fr; }}
      table {{ min-width: 980px; }}
      .table-wrap {{ overflow-x: auto; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>French OPA Dealings Tracker POC</h1>
    <div class="meta">
      <span>{len(rows)} parsed transaction rows</span>
      <span>{ok} notices clean</span>
      <span>{review} notices flagged for review</span>
      <span>Source: AMF BDIF DeclarationAchatVente</span>
    </div>
  </header>
  <main>
    <div class="toolbar">
      <label>Company
        <select id="company"><option value="">All companies</option></select>
      </label>
      <label>Buyer / seller
        <input id="filer" type="search" placeholder="Filter by filer">
      </label>
      <div>
        <button id="reset" type="button">Reset</button>
        <span class="count" id="count"></span>
      </div>
    </div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th style="width:90px">AMF ref</th>
            <th style="width:160px">Target</th>
            <th style="width:190px">Filer</th>
            <th style="width:110px">Date</th>
            <th style="width:110px">Deal</th>
            <th style="width:90px">Type</th>
            <th>Operation</th>
            <th style="width:100px">Quantity</th>
            <th style="width:90px">Price</th>
            <th style="width:220px">Resulting holding</th>
            <th style="width:70px">PDF</th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </div>
  </main>
  <script>
    const rows = {row_json};
    const company = document.querySelector("#company");
    const filer = document.querySelector("#filer");
    const tbody = document.querySelector("#rows");
    const count = document.querySelector("#count");

    const companies = [...new Set(rows.map(row => row.target).filter(Boolean))].sort();
    for (const name of companies) {{
      const option = document.createElement("option");
      option.value = name;
      option.textContent = name;
      company.append(option);
    }}

    function render() {{
      const companyValue = company.value;
      const filerValue = filer.value.trim().toLowerCase();
      const filtered = rows.filter(row => {{
        const companyOk = !companyValue || row.target === companyValue;
        const filerOk = !filerValue || String(row.filer || "").toLowerCase().includes(filerValue);
        return companyOk && filerOk;
      }});
      tbody.textContent = "";
      for (const row of filtered) {{
        const tr = document.createElement("tr");
        tr.innerHTML = `
          <td>${{escapeHtml(row.amf_number || "")}}</td>
          <td>${{escapeHtml(row.target || "")}}</td>
          <td>${{escapeHtml(row.filer || "")}}</td>
          <td>${{escapeHtml(row.transaction_date || "")}}</td>
          <td class="type-${{escapeHtml(row.operation_type || "other")}}">${{escapeHtml(row.operation_type || "")}}</td>
          <td>${{escapeHtml(row.instrument_type || "")}}</td>
          <td>${{escapeHtml(row.operation || "")}}</td>
          <td>${{escapeHtml(row.quantity || "")}}</td>
          <td>${{escapeHtml(row.price_eur || "")}}</td>
          <td>${{escapeHtml(row.resulting_holding || "")}}</td>
          <td>${{row.document_url ? `<a href="${{row.document_url}}">PDF</a>` : ""}}</td>
        `;
        tbody.append(tr);
      }}
      count.textContent = `${{filtered.length}} shown`;
    }}

    function escapeHtml(value) {{
      return String(value).replace(/[&<>"']/g, char => ({{
        "&": "&amp;",
        "<": "&lt;",
        ">": "&gt;",
        '"': "&quot;",
        "'": "&#39;"
      }}[char]));
    }}

    company.addEventListener("change", render);
    filer.addEventListener("input", render);
    document.querySelector("#reset").addEventListener("click", () => {{
      company.value = "";
      filer.value = "";
      render();
    }});
    render();
  </script>
</body>
</html>
"""
    path.write_text(content, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Parse AMF OPA DeclarationAchatVente PDFs.")
    parser.add_argument("--input", default="outputs/amf-france-poc/run/raw-filings.json")
    parser.add_argument("--out", default="outputs/amf-france-poc/run")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    raw = json.loads(input_path.read_text(encoding="utf-8"))
    base_dir = input_path.parent
    rows = []
    summaries = []

    for filing in raw.get("filings", []):
        local_pdf = filing.get("document", {}).get("local_pdf")
        if not local_pdf:
            continue
        pdf_path = base_dir / local_pdf
        text = extract_pdf_text(pdf_path)
        filing_rows, summary = parse_statement(filing, text)
        summary.update(
            {
                "amf_number": filing.get("amf_number"),
                "target": filing.get("target", {}).get("name"),
                "document_url": filing.get("document", {}).get("url"),
                "published_at": filing.get("published_at"),
                "online_at": filing.get("online_at"),
                "date_information": filing.get("date_information"),
            }
        )
        rows.extend(filing_rows)
        summaries.append(summary)

    offer_list_file = raw.get("source", {}).get("offer_period_company_list", {}).get("local_file")
    offer_companies = parse_offer_company_workbook(base_dir / offer_list_file) if offer_list_file else None

    parsed = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": raw.get("source"),
        "filter": raw.get("filter"),
        "notice_summaries": [
            {key: value for key, value in summary.items() if key != "raw_text"}
            for summary in summaries
        ],
        "offer_period_companies_from_amf_xlsx": offer_companies,
        "transactions": rows,
    }

    (out_dir / "filings.json").write_text(json.dumps(parsed, indent=2, ensure_ascii=False), encoding="utf-8")
    write_csv(out_dir / "filings.csv", rows)
    write_rss(out_dir / "feed.xml", raw.get("filings", []), rows)
    write_html(out_dir / "index.html", rows, summaries, raw.get("source"))

    print(
        json.dumps(
            {
                "notices": len(summaries),
                "transactions": len(rows),
                "review_notices": sum(1 for summary in summaries if summary["parse_status"] != "ok"),
                "out": str(out_dir),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
